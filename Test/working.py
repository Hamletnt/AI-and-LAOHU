import sys
from openpyxl import Workbook,load_workbook

# ตั้งค่า encoding ให้เป็น utf-8
sys.stdout.reconfigure(encoding='utf-8')

# โหลดไฟล์ workbook และ worksheet
wb = load_workbook('LProfitLoss24-Feb-.xlsx')
ws = wb.active

# พิมพ์ค่าในเซลล์ A1
print(ws['A3'].value)
