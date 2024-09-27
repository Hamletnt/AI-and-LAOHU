import sys
from openpyxl import load_workbook

# ตั้งค่า encoding ให้เป็น utf-8
sys.stdout.reconfigure(encoding='utf-8')

# โหลดไฟล์ workbook และ worksheet
wb = load_workbook('Test/LProfitLoss24-Feb-.xlsx')
ws = wb.active

# สร้าง dictionary เพื่อเก็บคอลัมน์ที่ตรงกับชื่อที่เราต้องการหา
column_mapping = {}

# รายการชื่อคอลัมน์ที่เราต้องการหา
keywords = ['วัตถุดิบ','หมู',' ไก่ ไข่','เนื้อ','วัชระ+สุรพล','ลูกชิ้น','กุ้ง','ผัก','ค่าไฟ','ไอติม','ขนมจีบ','ของหวาน','ผลไม้']

# วนลูปเพื่อหาชื่อคอลัมน์ในแถวที่ 1
for cell in ws[1]:  # อ่านทุกเซลล์ในแถวที่ 1
    if cell.value in keywords:  # ถ้าชื่อคอลัมน์ตรงกับ keyword
        column_mapping[cell.value] = cell.column_letter  # เก็บตัวอักษรคอลัมน์ไว้ใน dictionary

# แสดงผลคอลัมน์ที่ค้นหาได้
print(column_mapping)