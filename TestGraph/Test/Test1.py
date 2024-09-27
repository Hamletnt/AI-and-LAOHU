import sys
from openpyxl import load_workbook

# ตั้งค่า encoding ให้เป็น utf-8
sys.stdout.reconfigure(encoding='utf-8')

# โหลดไฟล์ workbook และ worksheet
wb = load_workbook('Test/LProfitLoss24-Feb-.xlsx',data_only=True)
ws = wb.active

# สร้าง array (list) เพื่อเก็บข้อมูลในคอลัมน์ C
column_c_data = []

# ตัวแปรเพื่อติดตามจำนวนครั้งที่เจอค่า None
none_count = 0

# เริ่มต้น loop ในคอลัมน์ C โดยเช็คว่าเจอค่า None หรือไม่
for row in ws.iter_rows(min_row=1, max_col=3, values_only=True):  # อ่านข้อมูลจากคอลัมน์ C
    cell_value = row[2]  # คอลัมน์ C คือ index 2
    
    if cell_value is None:
        none_count += 1  # นับจำนวนครั้งที่เจอ None
        if none_count >= 20:  # ถ้าเจอ None ติดต่อกัน 10 ครั้ง
            break  # หยุดการ loop
    else:
        none_count = 0  # รีเซ็ตจำนวนครั้งที่เจอ None ถ้าเจอค่าอื่น
    
    column_c_data.append(cell_value)

#ดูว่า total อยู่ที่บรรทัดไหน
count_row = len(column_c_data)

# สร้าง dictionary เพื่อเก็บคอลัมน์ที่ตรงกับชื่อที่เราต้องการหา
column_mapping = {}

# รายการชื่อคอลัมน์ที่เราต้องการหา
keywords = ['วัตถุดิบ','หมู',' ไก่ ไข่','เนื้อ','วัชระ+สุรพล','ลูกชิ้น','กุ้ง','ผัก','ค่าไฟ','ไอติม','ขนมจีบ','ของหวาน','ผลไม้']

# วนลูปเพื่อหาชื่อคอลัมน์ในแถวที่ 1
for cell in ws[1]:  # อ่านทุกเซลล์ในแถวที่ 1
    if cell.value in keywords:  # ถ้าชื่อคอลัมน์ตรงกับ keyword
        column_mapping[cell.value] = cell.column_letter  # เก็บตัวอักษรคอลัมน์ไว้ใน dictionary

#print(column_mapping)

# price_material = {
#     'วัตถุดิบอื่นๆ': ดึงข้อมูล column[วุตถุดิบ]แถวที่ count_row ,
#     'หมู': ,
#     'ไก่ไข่' : ,
#     'เนื้อ' : ,
#     'FarmF+สุรพล' ,
#     'ลูกชิ้น' : ,
#     'กุ้ง': ,
#     'ผัก': , 
#     'ค่านํ้า-ค่าไฟ' : , 
#     'ไอติม': ,
#     'ขนมจีบซาลาเปา': ,
#     'ของหวาน': ,
#     'ผลไม้': ,
# }

print(price_material)